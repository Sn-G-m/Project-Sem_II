# Function for adding attendance when face detected
def attendance(face, name):
    import cv2
    from openpyxl import load_workbook
    from openpyxl import Workbook
    from datetime import date

    today = date.today()
    date_string = today.strftime("%Y-%m-%d")
    # '2023-06-04'
    try:
        workbook = load_workbook('attendance.xlsx')
    except FileNotFoundError:
        workbook = Workbook()

    # Create a new sheet for the current month if it doesn't exist
        # today.strftime("%B") returns month
    if today.strftime("%B") not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=today.strftime("%B"))
    else:
        sheet = workbook[today.strftime("%B")]

        # Iterate over rows
    for row in sheet.iter_rows():
        # Iterate over all cells in the row
        for cell in row:
            # Check if the cell value is
            if cell.value == name:
                # Marking attendance for current date
                sheet[f"{cell.coordinate}"[0]+date_string[-2:].lstrip("0")] = 1
                cv2.putText(face, "marked", (20, 50),
                            cv2.FONT_HERSHEY_COMPLEX, 0.7, (255, 255, 0), 1)

                workbook.save('attendance.xlsx')
                workbook.close()

        else:
            pass

# To add new student


def add_name_to_excel_as_colnames(name):
    from openpyxl import load_workbook
    from openpyxl import Workbook
    from datetime import date
    # Create a excel file "attendance" if it doesn't exist
    try:
        workbook = load_workbook('attendance.xlsx')
    except FileNotFoundError:
        workbook = Workbook()

    today = date.today()

    # Create a sheet for the current month if it doesn't exist
    if today.strftime("%B") not in workbook.sheetnames:
        # today.strftime("%B") return name of current month
        sheet = workbook.create_sheet(title=today.strftime("%B"))
    else:
        sheet = workbook[today.strftime("%B")]

    cell_value = []
    for row in sheet.iter_rows():
        # Iterate over all cells in the row
        for cell in row:
            cell_value = cell_value + [cell.value]
    # Check if name is in data
    if name in cell_value:
        print("name already exists")
        workbook.close()
    else:
        # Find the next available column
        next_column = sheet.max_column + 1
        # Add the name to the next available column in row 1
        sheet.cell(row=1, column=next_column).value = name

    workbook.save('attendance.xlsx')
    workbook.close()
def manipulate(dict_paths, parent_folder):
    iterator_outside = 1
    import os
    import cv2
    for key in dict_paths:
        iterator_inside = 1
        image = cv2.imread(dict_paths[key])
        for i in range(2, 3):
            j = i / 2
            # Apply the brightness transformation
            transformed_image_b = cv2.convertScaleAbs(image, alpha=j)
            transformed_image_d = cv2.convertScaleAbs(image, alpha=1 / j)
            transformed_image_c = cv2.convertScaleAbs(image, beta=60 * j)
            file_name = os.path.basename(dict_paths[key])
            # Save the transformed images
            name_1 = "Bright" + str(i - 1) + file_name
            name_2 = "Dark" + str(i - 1) + file_name
            name_3 = "Contrast" + str(i - 1) + file_name
            print(name_1)
            print(dict_paths[key])
            # Path specifications
            transformed_image_path_1 = os.path.join(parent_folder, name_1)
            transformed_image_path_2 = os.path.join(parent_folder, name_2)
            transformed_image_path_3 = os.path.join(parent_folder, name_3)
            # Writing into the file
            cv2.imwrite(transformed_image_path_1, transformed_image_b)
            cv2.imwrite(transformed_image_path_2, transformed_image_d)
            cv2.imwrite(transformed_image_path_3, transformed_image_c)
            print("Iteration: ", iterator_outside, " ; ", iterator_inside)
            iterator_inside += 1
        iterator_outside += 1    
